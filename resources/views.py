from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.urls import reverse
from datetime import timedelta, date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import inch
import io
from .models import Resource, Skill, ResourceSkill, TimeEntry, ResourceAvailability
from .forms import ResourceForm, ResourceSkillFormSet, TimeEntryForm, ResourceAvailabilityForm, BulkTimeEntryForm

@login_required
def resource_list(request):
    """List all resources"""
    resources = Resource.objects.all()
    for resource in resources:
        resource.utilization = resource.current_utilization()
    
    return render(request, 'resources/resource_list.html', {'resources': resources})

@login_required
def resource_detail(request, pk):
    """Detail view for a resource"""
    resource = get_object_or_404(Resource, pk=pk)
    resource.utilization = resource.current_utilization()
    
    # Get recent time entries
    recent_time_entries = resource.time_entries.order_by('-date')[:10]
    
    # Get upcoming availability
    upcoming_availability = resource.availability.filter(
        start_date__gte=timezone.now().date()
    ).order_by('start_date')[:5]
    
    context = {
        'resource': resource,
        'recent_time_entries': recent_time_entries,
        'upcoming_availability': upcoming_availability
    }
    
    return render(request, 'resources/resource_detail.html', context)

@login_required
def resource_create(request):
    """Create a new resource"""
    if request.method == 'POST':
        form = ResourceForm(request.POST)
        formset = ResourceSkillFormSet(request.POST)
        
        if form.is_valid() and formset.is_valid():
            resource = form.save()
            formset.instance = resource
            formset.save()
            messages.success(request, 'Resource created successfully.')
            return redirect('resources:resource_detail', pk=resource.pk)
    else:
        form = ResourceForm()
        formset = ResourceSkillFormSet()
    
    return render(request, 'resources/resource_form.html', {
        'form': form,
        'formset': formset,
        'title': 'Create Resource'
    })

@login_required
def resource_edit(request, pk):
    """Edit an existing resource"""
    resource = get_object_or_404(Resource, pk=pk)
    
    if request.method == 'POST':
        form = ResourceForm(request.POST, instance=resource)
        formset = ResourceSkillFormSet(request.POST, instance=resource)
        
        if form.is_valid() and formset.is_valid():
            form.save()
            formset.save()
            messages.success(request, 'Resource updated successfully.')
            return redirect('resources:resource_detail', pk=resource.pk)
    else:
        form = ResourceForm(instance=resource)
        formset = ResourceSkillFormSet(instance=resource)
    
    return render(request, 'resources/resource_form.html', {
        'form': form,
        'formset': formset,
        'title': 'Edit Resource'
    })

@login_required
def create_skill(request):
    """Create a new skill via AJAX"""
    if request.method == 'POST' and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        skill_name = request.POST.get('skill_name')
        description = request.POST.get('description', '')
        
        if not skill_name:
            return JsonResponse({'success': False, 'error': 'Skill name is required'}, status=400)
            
        # Check if skill with this name already exists
        if Skill.objects.filter(name=skill_name).exists():
            return JsonResponse({'success': False, 'error': 'Skill with this name already exists'}, status=400)
            
        # Create new skill
        skill = Skill.objects.create(name=skill_name, description=description)
        
        # Return success with skill data
        return JsonResponse({
            'success': True,
            'skill': {
                'id': skill.id,
                'name': skill.name,
                'description': skill.description
            }
        })
    
    return JsonResponse({'success': False, 'error': 'Invalid request'}, status=400)

@login_required
def skill_list(request):
    """List all skills with option to delete them"""
    skills = Skill.objects.all()
    
    # Count resources using each skill for reference
    for skill in skills:
        skill.usage_count = ResourceSkill.objects.filter(skill=skill).count()
    
    return render(request, 'resources/skill_list.html', {'skills': skills})

@login_required
def skill_delete(request, pk):
    """Delete a skill"""
    skill = get_object_or_404(Skill, pk=pk)
    
    # Check if any resources are using this skill
    usage_count = ResourceSkill.objects.filter(skill=skill).count()
    
    if request.method == 'POST':
        skill_name = skill.name
        skill.delete()
        messages.success(request, f'Skill "{skill_name}" deleted successfully.')
        return redirect('skill_list')
    
    return render(request, 'resources/skill_confirm_delete.html', {
        'skill': skill,
        'usage_count': usage_count
    })

@login_required
def time_entry_list(request):
    """List time entries with filtering and summary statistics"""
    from django.db.models import Sum, Count, Q
    from projects.models import Project
    from decimal import Decimal
    
    # Start with all time entries
    time_entries = TimeEntry.objects.select_related('resource', 'task', 'task__project').order_by('-date')
    
    # Get filter parameters
    resource_id = request.GET.get('resource')
    project_id = request.GET.get('project')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    billable_filter = request.GET.get('billable')
    
    # Apply filters
    if resource_id:
        time_entries = time_entries.filter(resource_id=resource_id)
    
    if project_id:
        time_entries = time_entries.filter(task__project_id=project_id)
    
    if start_date:
        time_entries = time_entries.filter(date__gte=start_date)
    
    if end_date:
        time_entries = time_entries.filter(date__lte=end_date)
    
    if billable_filter:
        if billable_filter == 'true':
            time_entries = time_entries.filter(is_billable=True)
        elif billable_filter == 'false':
            time_entries = time_entries.filter(is_billable=False)
    
    # Calculate summary statistics
    summary_stats = time_entries.aggregate(
        total_entries=Count('id'),
        total_hours=Sum('hours')
    )
    
    # Calculate billable hours and statistics
    billable_entries = time_entries.filter(is_billable=True)
    billable_stats = billable_entries.aggregate(
        billable_hours=Sum('hours')
    )
    
    # Calculate totals with defaults
    total_entries = summary_stats['total_entries'] or 0
    total_hours = summary_stats['total_hours'] or Decimal('0')
    billable_hours = billable_stats['billable_hours'] or Decimal('0')
    
    # Calculate billable percentage
    billable_percentage = (billable_hours / total_hours * 100) if total_hours > 0 else 0
    
    # Calculate estimated value (simplified - using average cost per hour)
    estimated_value = Decimal('0')
    if billable_hours > 0:
        # Get average cost per hour from resources with billable time
        resources_with_billable_time = Resource.objects.filter(
            time_entries__in=billable_entries,
            cost_per_hour__isnull=False
        ).distinct()
        
        if resources_with_billable_time.exists():
            total_cost_per_hour = sum(r.cost_per_hour for r in resources_with_billable_time)
            avg_cost_per_hour = total_cost_per_hour / len(resources_with_billable_time)
            estimated_value = billable_hours * avg_cost_per_hour
    
    # Pagination - limit to 100 entries for performance
    time_entries = time_entries[:100]
    
    # Get dropdown data
    resources = Resource.objects.all().order_by('name')
    projects = Project.objects.all().order_by('name')
    
    context = {
        'time_entries': time_entries,
        'resources': resources,
        'projects': projects,
        'selected_resource': resource_id,
        'selected_project': project_id,
        'start_date': start_date,
        'end_date': end_date,
        'billable_filter': billable_filter,
        # Summary statistics
        'total_entries': total_entries,
        'total_hours': total_hours,
        'billable_hours': billable_hours,
        'billable_percentage': billable_percentage,
        'estimated_value': estimated_value,
    }
    
    return render(request, 'resources/time_entries.html', context)

@login_required
def time_entry_create(request):
    """Create a new time entry"""
    if request.method == 'POST':
        form = TimeEntryForm(request.POST, user=request.user)
        
        if form.is_valid():
            form.save()
            messages.success(request, 'Time entry recorded successfully.')
            return redirect('resources:time_entry_list')
    else:
        form = TimeEntryForm(user=request.user)
    
    return render(request, 'resources/time_entry_form.html', {
        'form': form,
        'title': 'Record Time Entry'
    })

@login_required
def time_entry_edit(request, pk):
    """Edit an existing time entry"""
    time_entry = get_object_or_404(TimeEntry, pk=pk)
    
    if request.method == 'POST':
        form = TimeEntryForm(request.POST, instance=time_entry, user=request.user)
        
        if form.is_valid():
            form.save()
            messages.success(request, 'Time entry updated successfully.')
            return redirect('resources:time_entry_list')
    else:
        form = TimeEntryForm(instance=time_entry, user=request.user)
    
    return render(request, 'resources/time_entry_form.html', {
        'form': form,
        'title': 'Edit Time Entry'
    })

@login_required
def time_entry_delete(request, pk):
    """Delete a time entry"""
    time_entry = get_object_or_404(TimeEntry, pk=pk)
    
    if request.method == 'POST':
        time_entry_info = f"{time_entry.resource.name} - {time_entry.date} ({time_entry.hours}h)"
        time_entry.delete()
        messages.success(request, f'Time entry "{time_entry_info}" deleted successfully.')
        return redirect('resources:time_entry_list')
    
    # For GET requests, redirect to edit page or time entry list
    # Since there's no dedicated delete confirmation template, just delete on POST
    return redirect('resources:time_entry_list')

@login_required
def bulk_time_entry(request):
    """Create multiple time entries for a date range"""
    if request.method == 'POST':
        form = BulkTimeEntryForm(request.POST, user=request.user)
        
        if form.is_valid():
            # Create time entries for each day in the range
            resource = form.cleaned_data['resource']            
            task = form.cleaned_data['task']
            start_date = form.cleaned_data['start_date']
            end_date = form.cleaned_data['end_date']
            hours_per_day = form.cleaned_data['hours_per_day']
            description = form.cleaned_data['description']
            is_billable = form.cleaned_data.get('is_billable', True)
            include_weekends = form.cleaned_data.get('include_weekends', False)
            created_count = 0
            updated_count = 0
            current_date = start_date
            
            while current_date <= end_date:
                # Skip weekends unless specifically included
                should_create = True
                if not include_weekends and current_date.weekday() >= 5:  # Saturday (5) or Sunday (6)
                    should_create = False
                
                if should_create:
                    time_entry, created = TimeEntry.objects.get_or_create(
                        resource=resource,
                        task=task,
                        date=current_date,                        
                        defaults={
                            'hours': hours_per_day,
                            'description': description,
                            'is_billable': is_billable
                        }
                    )
                    
                    if created:
                        created_count += 1
                    else:
                        # Entry already exists, update it
                        time_entry.hours = hours_per_day
                        time_entry.description = description
                        time_entry.is_billable = is_billable
                        time_entry.save()
                        updated_count += 1
                        
                current_date += timedelta(days=1)
            
            # Create appropriate success message
            if created_count > 0 and updated_count > 0:
                message = f'Created {created_count} new time entries and updated {updated_count} existing entries.'
            elif created_count > 0:
                message = f'Created {created_count} time entries.'
            elif updated_count > 0:
                message = f'Updated {updated_count} existing time entries.'
            else:
                message = 'No time entries were created or updated.'
            
            messages.success(request, message)
            return redirect('resources:time_entry_list')
    else:
        form = BulkTimeEntryForm(user=request.user)
    
    return render(request, 'resources/bulk_time_entry_form.html', {
        'form': form,
        'title': 'Bulk Time Entry'
    })

@login_required
def bulk_time_action(request):
    """Handle bulk actions on time entries (delete, etc.)"""
    if request.method != 'POST':
        return redirect('resources:time_entry_list')
    
    action = request.POST.get('action')
    entry_ids = request.POST.getlist('entry_ids')
    
    if not action or not entry_ids:
        messages.error(request, 'No action or entries selected.')
        return redirect('resources:time_entry_list')
    
    # Convert string IDs to integers
    try:
        entry_ids = [int(id_) for id_ in entry_ids]
    except ValueError:
        messages.error(request, 'Invalid entry IDs.')
        return redirect('resources:time_entry_list')
      # Get the time entries to operate on
    time_entries = TimeEntry.objects.filter(id__in=entry_ids)
    
    if action == 'delete':
        count = time_entries.count()
        time_entries.delete()
        messages.success(request, f'Successfully deleted {count} time entries.')
    elif action == 'mark-billable':
        count = time_entries.update(is_billable=True)
        messages.success(request, f'Successfully marked {count} time entries as billable.')
    elif action == 'mark-non-billable':
        count = time_entries.update(is_billable=False)
        messages.success(request, f'Successfully marked {count} time entries as non-billable.')
    else:
        messages.error(request, f'Unknown action: {action}')
    
    return redirect('resources:time_entry_list')

@login_required
def availability_calendar(request):
    """Display availability calendar"""
    # Handle POST request for creating availability entries
    if request.method == 'POST':
        form = ResourceAvailabilityForm(request.POST, user=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Availability entry created successfully.')
            # Redirect to avoid form resubmission
            resource_id = request.GET.get('resource')
            if resource_id:
                return redirect(f"{reverse('resources:availability_calendar')}?resource={resource_id}")
            return redirect('resources:availability_calendar')
    else:
        form = ResourceAvailabilityForm(user=request.user)
    
    resource_id = request.GET.get('resource')
    
    # Get availability data for calendar display (broader range for calendar view)
    start_date = timezone.now().date().replace(day=1)  # Start of current month
    end_date = (start_date + timedelta(days=90)).replace(day=1) - timedelta(days=1)  # ~3 months ahead
    
    calendar_events = ResourceAvailability.objects.filter(
        start_date__lte=end_date,
        end_date__gte=start_date
    ).select_related('resource')
    
    # Get upcoming events for the sidebar (next 30 days)
    upcoming_start = timezone.now().date()
    upcoming_end = upcoming_start + timedelta(days=30)
    
    upcoming_events = ResourceAvailability.objects.filter(
        start_date__lte=upcoming_end,
        end_date__gte=upcoming_start
    ).select_related('resource').order_by('start_date')
    
    # Filter by resource if specified
    if resource_id:
        calendar_events = calendar_events.filter(resource_id=resource_id)
        upcoming_events = upcoming_events.filter(resource_id=resource_id)
    
    resources = Resource.objects.all()
    
    context = {
        'calendar_events': calendar_events,
        'upcoming_events': upcoming_events,
        'availability_form': form,
        'resources': resources,
        'selected_resource': int(resource_id) if resource_id else None,
        'start_date': start_date,
        'end_date': end_date
    }
    
    return render(request, 'resources/availability_calendar.html', context)

@login_required
def availability_create(request):
    """Create availability entry"""
    if request.method == 'POST':
        form = ResourceAvailabilityForm(request.POST, user=request.user)
        
        if form.is_valid():
            form.save()
            messages.success(request, 'Availability updated successfully.')
            return redirect('availability_calendar')
    else:
        form = ResourceAvailabilityForm(user=request.user)
    
    return render(request, 'resources/availability_form.html', {
        'form': form,
        'title': 'Add Availability'
    })

@login_required
def availability_edit(request, pk):
    """Edit availability entry"""
    availability = get_object_or_404(ResourceAvailability, pk=pk)
    
    if request.method == 'POST':
        form = ResourceAvailabilityForm(request.POST, instance=availability)
        
        if form.is_valid():
            form.save()
            messages.success(request, 'Availability updated successfully.')
            return redirect('availability_calendar')
    else:
        form = ResourceAvailabilityForm(instance=availability)
    
    return render(request, 'resources/availability_form.html', {
        'form': form,
        'title': 'Edit Availability'
    })

@login_required
def resource_time_tracking_report(request, pk):
    """Generate time tracking report for a specific resource"""
    resource = get_object_or_404(Resource, pk=pk)
    
    # Get date range from query parameters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    # Default to current month if no dates provided
    if not start_date or not end_date:
        today = timezone.now().date()
        start_date = today.replace(day=1)
        if today.month == 12:
            end_date = today.replace(year=today.year + 1, month=1, day=1) - timedelta(days=1)
        else:
            end_date = today.replace(month=today.month + 1, day=1) - timedelta(days=1)
    else:
        try:
            start_date = timezone.datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date = timezone.datetime.strptime(end_date, '%Y-%m-%d').date()
        except ValueError:
            messages.error(request, 'Invalid date format. Using current month.')
            today = timezone.now().date()
            start_date = today.replace(day=1)
            if today.month == 12:
                end_date = today.replace(year=today.year + 1, month=1, day=1) - timedelta(days=1)
            else:
                end_date = today.replace(month=today.month + 1, day=1) - timedelta(days=1)
    
    # Get time entries for the resource in the date range
    time_entries = resource.time_entries.filter(
        date__gte=start_date,
        date__lte=end_date
    ).select_related('task', 'task__project').order_by('-date')
    
    # Calculate summary statistics
    total_hours = sum(entry.hours for entry in time_entries)
    total_days = len(set(entry.date for entry in time_entries))
    
    # Group by project
    project_hours = {}
    for entry in time_entries:
        project_name = entry.task.project.name
        if project_name not in project_hours:
            project_hours[project_name] = 0
        project_hours[project_name] += float(entry.hours)
    
    context = {
        'resource': resource,
        'time_entries': time_entries,
        'start_date': start_date,
        'end_date': end_date,
        'total_hours': total_hours,
        'total_days': total_days,
        'project_hours': project_hours,
        'average_hours_per_day': round(total_hours / total_days, 2) if total_days > 0 else 0,
    }
    
    return render(request, 'resources/time_tracking_report.html', context)

@login_required
def export_time_entries(request):
    """Export time entries to PDF or Excel"""
    format_type = request.GET.get('format', 'pdf')
    
    # Get filter parameters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    project_id = request.GET.get('project')
    resource_id = request.GET.get('resource')
    billable_filter = request.GET.get('billable')
    
    # Build queryset with filters
    queryset = TimeEntry.objects.select_related('resource', 'task__project').order_by('-date')
    
    if start_date:
        queryset = queryset.filter(date__gte=start_date)
    if end_date:
        queryset = queryset.filter(date__lte=end_date)
    if project_id and project_id != 'all':
        queryset = queryset.filter(task__project_id=project_id)
    if resource_id and resource_id != 'all':
        queryset = queryset.filter(resource_id=resource_id)
    if billable_filter and billable_filter != 'all':
        if billable_filter == 'billable':
            queryset = queryset.filter(is_billable=True)
        elif billable_filter == 'non_billable':
            queryset = queryset.filter(is_billable=False)
    
    time_entries = list(queryset)
    
    if format_type == 'excel':
        return _export_time_entries_excel(time_entries, start_date, end_date)
    else:
        return _export_time_entries_pdf(time_entries, start_date, end_date)

def _export_time_entries_excel(time_entries, start_date, end_date):
    """Export time entries to Excel format"""
    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Time Entries"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Add title
    title = f"Time Entries Report"
    if start_date and end_date:
        title += f" ({start_date} to {end_date})"
    
    ws.merge_cells('A1:H1')
    ws['A1'] = title
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal="center")
    
    # Add summary
    total_entries = len(time_entries)
    total_hours = sum(float(entry.hours) for entry in time_entries)
    billable_hours = sum(float(entry.hours) for entry in time_entries if entry.is_billable)
    
    ws['A3'] = f"Total Entries: {total_entries}"
    ws['C3'] = f"Total Hours: {total_hours:.2f}"
    ws['E3'] = f"Billable Hours: {billable_hours:.2f}"
    ws['G3'] = f"Billable Rate: {(billable_hours/total_hours*100):.1f}%" if total_hours > 0 else "0%"
    
    # Headers
    headers = ['Date', 'Resource', 'Project', 'Task', 'Hours', 'Description', 'Billable', 'Value']
    row = 5
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Data rows
    for entry in time_entries:
        row += 1
        ws.cell(row=row, column=1, value=entry.date.strftime('%Y-%m-%d'))
        ws.cell(row=row, column=2, value=entry.resource.name)
        ws.cell(row=row, column=3, value=entry.task.project.name)
        ws.cell(row=row, column=4, value=entry.task.name)
        ws.cell(row=row, column=5, value=float(entry.hours))
        ws.cell(row=row, column=6, value=entry.description or '')
        ws.cell(row=row, column=7, value='Yes' if entry.is_billable else 'No')
        # Estimate value (you can customize this calculation)
        estimated_value = float(entry.hours) * 100 if entry.is_billable else 0  # $100/hour example rate
        ws.cell(row=row, column=8, value=estimated_value)    # Auto-adjust column widths (safer approach)
    column_widths = [15, 20, 25, 30, 10, 30, 10, 15]  # Predefined widths for each column
    for i, width in enumerate(column_widths, 1):
        column_letter = get_column_letter(i)
        ws.column_dimensions[column_letter].width = width
    
    # Save to response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="time_entries_export.xlsx"'
    
    # Save workbook to response
    wb.save(response)
    return response

def _export_time_entries_pdf(time_entries, start_date, end_date):
    """Export time entries to PDF format"""
    # Create response
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="time_entries_export.pdf"'
    
    # Create PDF document
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch)
    
    # Get styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1  # Center alignment
    )
    
    # Build content
    content = []
    
    # Title
    title = f"Time Entries Report"
    if start_date and end_date:
        title += f" ({start_date} to {end_date})"
    content.append(Paragraph(title, title_style))
    
    # Summary
    total_entries = len(time_entries)
    total_hours = sum(float(entry.hours) for entry in time_entries)
    billable_hours = sum(float(entry.hours) for entry in time_entries if entry.is_billable)
    billable_rate = (billable_hours/total_hours*100) if total_hours > 0 else 0
    
    summary_data = [
        ['Total Entries:', str(total_entries), 'Total Hours:', f"{total_hours:.2f}"],
        ['Billable Hours:', f"{billable_hours:.2f}", 'Billable Rate:', f"{billable_rate:.1f}%"]
    ]
    
    summary_table = Table(summary_data, colWidths=[1.5*inch, 1*inch, 1.5*inch, 1*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.lightgrey),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    content.append(summary_table)
    content.append(Spacer(1, 20))
    
    # Time entries table
    if time_entries:
        # Table headers
        table_data = [['Date', 'Resource', 'Project', 'Task', 'Hours', 'Billable']]
        
        # Table rows
        for entry in time_entries:
            table_data.append([
                entry.date.strftime('%Y-%m-%d'),
                entry.resource.name,
                entry.task.project.name[:20] + '...' if len(entry.task.project.name) > 20 else entry.task.project.name,
                entry.task.name[:25] + '...' if len(entry.task.name) > 25 else entry.task.name,
                f"{entry.hours}h",
                'Yes' if entry.is_billable else 'No'
            ])
        
        # Create table
        table = Table(table_data, colWidths=[1*inch, 1.5*inch, 1.8*inch, 2*inch, 0.8*inch, 0.8*inch])
        table.setStyle(TableStyle([
            # Header style
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            
            # Data style
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            
            # Alternating row colors
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
        ]))
        content.append(table)
    else:
        content.append(Paragraph("No time entries found.", styles['Normal']))
    
    # Build PDF
    doc.build(content)
    
    # Get PDF data
    pdf_data = buffer.getvalue()
    buffer.close()
    
    response.write(pdf_data)
    return response

@login_required
def toggle_time_entry_billable(request, pk):
    """Toggle the billable status of a time entry"""
    time_entry = get_object_or_404(TimeEntry, pk=pk)
    
    if request.method == 'POST':
        # Toggle the billable status
        time_entry.is_billable = not time_entry.is_billable
        time_entry.save()
        
        status = "billable" if time_entry.is_billable else "non-billable"
        messages.success(request, f'Time entry marked as {status}.')
    
    return redirect('resources:time_entry_list')
